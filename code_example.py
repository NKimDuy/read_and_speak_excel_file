import io
import sounddevice as sd
import soundfile as sf
from gtts import gTTS
from openpyxl import load_workbook


def preprocess(text):
    # Tách từng ký tự, đọc rời nhau
    result = []
    for char in text:
        if char.isalpha():
            result.append(char.upper())  # A, B, C...
        elif char.isdigit():
            result.append(char)          # 9, 4, 2, 3...
        else:
            if char == "-":
                result.append("gạch")     
            if char == "/":
                result.append("xuyệt")
            if char == ".":
                result.append("chấm")
            if char == ",":
                result.append("phẩy")
    return " ".join(result)


def read_and_speak(text):
    tts = gTTS(text=text, lang='vi')
    buf = io.BytesIO()
    tts.write_to_fp(buf)
    buf.seek(0)
    data, samplerate = sf.read(buf)
    sd.play(data, samplerate)
    sd.wait()
 
# Load the Excel file
workbook = load_workbook("test.xlsx")
sheet = workbook.active
headers = [cell.value for cell in sheet[1]]

for row in sheet.iter_rows(min_row=2, values_only=True):  
    ho = row[0]  
    ten = row[1]
    ngay_sinh = preprocess(row[2].strftime("%d/%m/%Y"))
 
    read_and_speak(headers[0])
    read_and_speak(ho)
    read_and_speak(headers[1])
    read_and_speak(ten)
    read_and_speak(headers[2])
    read_and_speak(ngay_sinh)

