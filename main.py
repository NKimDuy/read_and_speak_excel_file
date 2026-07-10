import threading
import tkinter as tk
from tkinter import filedialog, ttk

import sounddevice as sd
import soundfile as sf
from gtts import gTTS
from openpyxl import load_workbook
import librosa
import numpy as np
import tempfile
import os

# ============================================================
# PHẦN 1 — XỬ LÝ TEXT TRƯỚC KHI ĐỌC
# ============================================================

# Một số chữ cái gTTS đọc rời rất dễ nghe nhầm với chữ khác
# (VD: B với P, Đ với D...) nên phải đánh vần rõ thay vì đọc mỗi chữ cái.
# Muốn đổi cách đọc chữ nào thì sửa trực tiếp ở đây.
AMBIGUOUS_LETTERS = {
    "Đ": "đờ",
    "P": "bê phở",
    "B": "bê bò",
    "G": "rờ",
}

# Ký tự đặc biệt → đọc thành chữ tiếng Việt tương ứng
SYMBOL_TO_WORD = {
    "-": "gạch",
    "/": "xuyệt",
    ".": "chấm",
    ",": "phẩy",
}


def preprocess(text):
    """
    Chuyển text thô thành dạng gTTS có thể đọc tiếng Việt đúng.
    - Chữ cái: đọc rời (gTTS tự đọc), trừ các chữ dễ nghe nhầm ở AMBIGUOUS_LETTERS
    - Số: giữ nguyên, đọc rời từng chữ số
    - Ký tự đặc biệt (- / . ,): đổi thành chữ tương ứng
    - Ký tự khác: bỏ qua
    """
    if text is None:
        return ""  # Tránh lỗi nếu ô Excel bị trống

    text = str(text)  # Đảm bảo luôn là string, kể cả số hay datetime

    result = []
    for char in text:
        if char.isalpha():
            result.append(AMBIGUOUS_LETTERS.get(char, char))  # Chữ cái đọc rời nhau: A, B, C...
        elif char.isdigit():
            result.append(char)  # Số đọc rời nhau: 1, 9, 9, 0...
        elif char in SYMBOL_TO_WORD:
            result.append(SYMBOL_TO_WORD[char])
        elif char == " ":
            result.append(" ")  # Giữ khoảng trắng để ngắt từ tự nhiên
        # Các ký tự khác bỏ qua luôn

    return " ".join(result)


# ============================================================
# PHẦN 2 — ỨNG DỤNG TKINTER
# ============================================================

class ExcelTTSApp:
    """
    Class chứa toàn bộ logic của app.
    Dùng class để các phần của app có thể chia sẻ dữ liệu với nhau
    thông qua self (ví dụ self.rows, self.current_index...)
    """

    # --- CẤU HÌNH CỘT — SỬA LẠI CHO KHỚP NẾU CẤU TRÚC FILE EXCEL THAY ĐỔI ---
    # Các cột cần đánh vần từng chữ/số (index tính từ 0, VD: mã hồ sơ, số CMND/CCCD...)
    SPELL_OUT_COLUMNS = (0, 3, 7, 8)
    # Cột "ngành" — nếu trùng với dòng ngay trước thì chỉ đọc "cùng ngành" cho đỡ lặp
    SAME_AS_PREVIOUS_COLUMN = 6

    def __init__(self, root):
        """
        Hàm khởi tạo — chạy một lần khi app mở.
        root: cửa sổ Tkinter chính
        """
        self.root = root
        self.root.title("Excel TTS Reader")
        self.root.geometry("900x600")
        self.root.configure(bg="#1e1e2e")  # Màu nền tối

        # --- DỮ LIỆU ---
        self.rows = []          # List các tuple — toàn bộ dữ liệu Excel
        self.headers = []       # List tên cột
        self.workbook = None    # Object workbook để ghi X vào file
        self.filepath = None    # Đường dẫn file Excel đang mở

        self.current_index = 0  # Index dòng đang đọc (tính từ 0)
        self.page = 0           # Trang hiện tại (tính từ 0)
        self.page_size = 20     # Số dòng mỗi trang

        # --- TRẠNG THÁI ---
        # is_reading: app có đang đọc không
        self.is_reading = False

        # stop_requested: người dùng đã nhấn Space chưa
        # Nếu True thì đọc xong dòng hiện tại sẽ dừng
        self.stop_requested = False

        # reading_thread: thread đang chạy TTS
        # Cần lưu lại để kiểm tra thread còn sống không
        self.reading_thread = None

        # Kiểm tra thông tin ngành dòng trên đã có rồi
        # thì dòng dưới không cần đọc nữa
        self.previous_row = None

        # mặc định tốc độ đọc trung bình
        self.speed_var = tk.DoubleVar(value=1.0)  # mặc định 1.0

        self.audio_lock = threading.Lock()   # khoá bảo vệ thiết bị âm thanh

        # --- GIAO DIỆN ---
        self._build_ui()

        # Bind phím — root.bind để lắng nghe phím từ bất kỳ đâu trong app
        self.root.bind("<Return>", self.on_enter)   # Enter
        self.root.bind("<space>", self.on_space)    # Space

    def _build_ui(self):
        """
        Tạo toàn bộ giao diện.
        Tách ra hàm riêng để __init__ không quá dài.
        """

        # --- THANH TRÊN CÙNG: nút mở file + trạng thái ---
        top_frame = tk.Frame(self.root, bg="#1e1e2e")
        top_frame.pack(fill="x", padx=16, pady=(16, 8))

        btn_open = tk.Button(
            top_frame,
            text="📂 Mở file Excel",
            command=self.open_file,         # Gọi hàm open_file khi click
            bg="#7c3aed", fg="white",
            font=("Consolas", 11, "bold"),
            relief="flat", padx=12, pady=6,
            cursor="hand2"
        )
        btn_open.pack(side="left")

        # Label hiện trạng thái (đang đọc / dừng / chưa mở file)
        self.status_label = tk.Label(
            top_frame,
            text="Chưa mở file",
            bg="#1e1e2e", fg="#a0a0b0",
            font=("Consolas", 10)
        )
        self.status_label.pack(side="left", padx=16)

        # Label hiện tiến độ (dòng X / tổng Y)
        self.progress_label = tk.Label(
            top_frame,
            text="",
            bg="#1e1e2e", fg="#7c3aed",
            font=("Consolas", 10, "bold")
        )
        self.progress_label.pack(side="right")

        # --- HƯỚNG DẪN PHÍM ---
        hint_frame = tk.Frame(self.root, bg="#1e1e2e")
        hint_frame.pack(fill="x", padx=16, pady=(0, 8))

        tk.Label(
            hint_frame,
            text="Enter: bắt đầu/tiếp tục   |   Space: dừng sau dòng hiện tại",
            bg="#1e1e2e", fg="#555570",
            font=("Consolas", 9)
        ).pack(side="left")

        # Radio button tốc độ — bên phải hint_frame
        tk.Label(
            hint_frame,
            text=": Tốc độ",
            bg="#1e1e2e", fg="#a0a0b0",
            font=("Consolas", 9)
        ).pack(side="right", padx=(8, 4))

        # Label hiện số tốc độ hiện tại
        self.speed_label = tk.Label(
            hint_frame,
            text="1.0x",
            bg="#1e1e2e", fg="#7c3aed",
            font=("Consolas", 9, "bold"),
            width=4
        )
        self.speed_label.pack(side="right")

        # Thanh kéo — từ 0.5x đến 2.0x, bước 0.1
        tk.Scale(
            hint_frame,
            from_=0.5, to=2.0,        # min và max
            resolution=0.1,            # bước nhảy mỗi lần kéo
            orient="horizontal",       # kéo ngang
            variable=self.speed_var,   # gắn với DoubleVar
            command=self._on_speed_change,  # gọi khi kéo
            bg="#1e1e2e", fg="#a0a0b0",
            highlightthickness=0,      # bỏ viền
            troughcolor="#2a2a3e",     # màu rãnh
            activebackground="#7c3aed",
            length=120,                # chiều dài thanh kéo (px)
            showvalue=False            # ẩn số mặc định của Scale, dùng label riêng
        ).pack(side="right")


        # --- BẢNG DỮ LIỆU ---
        # Frame chứa bảng + scrollbar dọc
        table_frame = tk.Frame(self.root, bg="#1e1e2e")
        table_frame.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        # Scrollbar dọc — cần tạo trước để truyền vào Treeview
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical")
        scrollbar.pack(side="right", fill="y")

        # Treeview là widget dạng bảng của Tkinter
        # show="headings" để ẩn cột tree mặc định (cột ngoài cùng bên trái)
        self.tree = ttk.Treeview(
            table_frame,
            yscrollcommand=scrollbar.set,  # Kết nối scrollbar với bảng
            show="headings",
            selectmode="none"              # Không cho chọn dòng bằng chuột
        )
        self.tree.pack(fill="both", expand=True)

        # Kết nối scrollbar ngược lại với bảng
        scrollbar.config(command=self.tree.yview)

        # Style cho Treeview — màu nền tối
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Treeview",
            background="#12121e",
            foreground="#d0d0e0",
            fieldbackground="#12121e",
            rowheight=32,
            font=("Consolas", 10)
        )
        style.configure(
            "Treeview.Heading",
            background="#2a2a3e",
            foreground="#7c3aed",
            font=("Consolas", 10, "bold")
        )

        # Tag màu cho các trạng thái dòng
        # Tag "reading": dòng đang đọc → nền vàng
        self.tree.tag_configure("reading", background="#3d3500", foreground="#ffd700")
        # Tag "done": dòng đã đọc xong → nền xanh
        self.tree.tag_configure("done", background="#0d3320", foreground="#4ade80")
        # Tag "normal": dòng chưa đọc → màu mặc định
        self.tree.tag_configure("normal", background="#12121e", foreground="#d0d0e0")

        # --- THANH PHÂN TRANG ---
        page_frame = tk.Frame(self.root, bg="#1e1e2e")
        page_frame.pack(fill="x", padx=16, pady=(0, 16))

        self.btn_prev = tk.Button(
            page_frame,
            text="◀ Trang trước",
            command=self.prev_page,
            bg="#2a2a3e", fg="#a0a0b0",
            font=("Consolas", 10),
            relief="flat", padx=10, pady=4,
            cursor="hand2"
        )
        self.btn_prev.pack(side="left")

        self.page_label = tk.Label(
            page_frame,
            text="",
            bg="#1e1e2e", fg="#a0a0b0",
            font=("Consolas", 10)
        )
        self.page_label.pack(side="left", padx=12)

        self.btn_next = tk.Button(
            page_frame,
            text="Trang sau ▶",
            command=self.next_page,
            bg="#2a2a3e", fg="#a0a0b0",
            font=("Consolas", 10),
            relief="flat", padx=10, pady=4,
            cursor="hand2"
        )
        self.btn_next.pack(side="left")

    # ============================================================
    # PHẦN 4 — MỞ FILE EXCEL
    # ============================================================

    def _on_speed_change(self, value=None): 
        rate = float(value)
        self.speed_label.config(text=f"{rate:.1f}x")  # hiện số như "1.2x"

    def _find_resume_index(self):
        """
        Quét cột cuối + 1 của file Excel để tìm dòng đầu tiên chưa có X.
        Trả về index đó để current_index bắt đầu từ đây.
        """
        sheet = self.workbook.active
        last_col = len(self.headers) + 1  # Cột chứa dấu X — cột cuối + 1

        for i, row in enumerate(sheet.iter_rows(min_row=2, min_col=last_col, max_col=last_col, values_only=True)):
            # row là tuple 1 phần tử vì chỉ đọc 1 cột — lấy phần tử đầu
            value = row[0]
            if value != "X":
                return i  # Dòng đầu tiên chưa có X → bắt đầu từ đây

        # Tất cả đều có X → đã đọc hết file
        return len(self.rows)
  

    def speak(self, text):
        """
        Nhận text, dùng gTTS tạo audio trong RAM rồi phát ra loa.
        Không tạo file trên ổ cứng.
        """
        tts = gTTS(text=text, lang='vi')
        tmp_path = tempfile.mktemp(suffix=".mp3")
        tts.save(tmp_path)
        data, samplerate = sf.read(tmp_path)
        os.remove(tmp_path)

        # Chụp tốc độ ra biến riêng NGAY ĐÂY — tránh giá trị đổi giữa chừng
        speed = self.speed_var.get()
        if abs(speed - 1.0) > 0.01:
            if data.ndim > 1:
                data = data.mean(axis=1)
            data = librosa.effects.time_stretch(data.astype(np.float32), rate=speed)

        # Bọc trong khoá — đảm bảo phát xong hẳn mới nhả cho thao tác khác
        with self.audio_lock:
            sd.stop()              # dừng hẳn stream cũ nếu còn sót
            sd.play(data, samplerate)
            sd.wait()

    def speak_row(self, row_data, headers):
        """
        Đọc toàn bộ một dòng Excel theo format: [tiêu đề]: [giá trị]
        row_data: tuple các giá trị của dòng đó
        headers: list tiêu đề cột
        """
        for i, value in enumerate(row_data):
            if value is None:
                continue  # Bỏ qua ô trống

            if i in self.SPELL_OUT_COLUMNS:
                processed = preprocess(value)
                self.speak(processed)
            elif i == self.SAME_AS_PREVIOUS_COLUMN:
                if self.previous_row is not None and self.previous_row[i] == value:
                    # Giống dòng trước → đọc "cùng ngành" thay vì lặp lại
                    self.speak("cùng ngành")
                else:
                    self.speak(str(value))
            else:
                self.speak(str(value))

        # Lưu dòng này để so sánh với dòng tiếp theo
        self.previous_row = row_data

    def open_file(self):
        """
        Mở hộp thoại chọn file Excel, đọc dữ liệu vào self.rows.
        """
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not filepath:
            return  # Người dùng bấm Cancel, không làm gì

        self.filepath = filepath

        # Load workbook với data_only=True để lấy giá trị ô,
        # không lấy công thức
        self.workbook = load_workbook(filepath, data_only=True)
        sheet = self.workbook.active

        # Lấy tiêu đề từ dòng đầu tiên
        self.headers = [cell.value for cell in sheet[1] if cell.value is not None]

        # Đọc toàn bộ dữ liệu từ dòng 2 trở đi vào RAM
        # values_only=True: lấy giá trị, không lấy object Cell
        self.rows = list(sheet.iter_rows(min_row=2, values_only=True))

        # Reset về trạng thái ban đầu, tự tìm dòng cần đọc tiếp (dựa vào cột đánh dấu X)
        self.current_index = self._find_resume_index()
        self.page = 0
        self.is_reading = False
        self.stop_requested = False
        # reset dòng dùng để so sánh đọc tên ngành
        self.previous_row = None

        # Cập nhật giao diện
        self._setup_table_columns()
        self._render_page()
        self.status_label.config(text="Sẵn sàng — nhấn Enter để bắt đầu")
        self._update_progress()

    def _setup_table_columns(self):
        """
        Tạo các cột cho Treeview dựa trên tiêu đề file Excel.
        Cần gọi lại mỗi khi mở file mới vì số cột có thể khác nhau.
        """
        # Xóa cột cũ nếu có
        self.tree["columns"] = self.headers
        for col in self.headers:
            self.tree.heading(col, text=col)
            # Chia đều chiều rộng — 860px / số cột
            self.tree.column(col, width=860 // len(self.headers), anchor="w")

    # ============================================================
    # PHẦN 5 — PHÂN TRANG VÀ RENDER
    # ============================================================

    def _render_page(self):
        """
        Xóa bảng và vẽ lại 20 dòng của trang hiện tại.
        Gọi mỗi khi chuyển trang hoặc cần cập nhật màu dòng.
        """
        # Xóa toàn bộ dòng hiện tại trong bảng
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Tính index bắt đầu và kết thúc của trang hiện tại
        start = self.page * self.page_size           # VD trang 0: start=0
        end = start + self.page_size                 # VD trang 0: end=20

        # Lấy 20 dòng của trang này
        page_rows = self.rows[start:end]

        for i, row in enumerate(page_rows):
            actual_index = start + i  # Index thực trong self.rows

            # Chuyển None thành chuỗi rỗng để hiện trong bảng
            display_values = [
                str(v) if v is not None else "" for v in row
            ]

            # Xác định tag màu cho dòng này
            if actual_index == self.current_index and self.is_reading:
                tag = "reading"   # Đang đọc → vàng
            elif actual_index < self.current_index:
                tag = "done"      # Đã đọc → xanh
            else:
                tag = "normal"    # Chưa đọc → mặc định

            self.tree.insert("", "end", values=display_values, tags=(tag,))

        # Cập nhật label trang
        total_pages = max(1, (len(self.rows) + self.page_size - 1) // self.page_size)
        self.page_label.config(text=f"Trang {self.page + 1} / {total_pages}")

    def _go_to_page_of(self, index):
        """
        Chuyển đến trang chứa dòng có index cho trước.
        Dùng khi đang đọc tự động để trang follow theo dòng đang đọc.
        """
        target_page = index // self.page_size
        if target_page != self.page:
            self.page = target_page
            # after(0, ...) để chạy trên main thread —
            # vì _render_page cập nhật UI, không được gọi từ thread phụ
            self.root.after(0, self._render_page)

    def prev_page(self):
        """Chuyển về trang trước — chỉ hoạt động khi đang dừng."""
        if self.is_reading:
            return  # Không cho chuyển trang khi đang đọc
        if self.page > 0:
            self.page -= 1
            self._render_page()

    def next_page(self):
        """Chuyển sang trang sau — chỉ hoạt động khi đang dừng."""
        if self.is_reading:
            return
        total_pages = (len(self.rows) + self.page_size - 1) // self.page_size
        if self.page < total_pages - 1:
            self.page += 1
            self._render_page()

    # ============================================================
    # PHẦN 6 — ĐIỀU KHIỂN ĐỌC
    # ============================================================

    def on_enter(self, event=None):
        """
        Xử lý phím Enter.
        - Nếu chưa có file → bỏ qua
        - Nếu đang đọc → bỏ qua (không restart)
        - Nếu đang dừng → bắt đầu/tiếp tục đọc
        """
        if not self.rows:
            return
        if self.is_reading:
            return

        # Reset cờ dừng — phòng trường hợp trước đó đã nhấn Space
        self.stop_requested = False
        self.is_reading = True
        self.status_label.config(text="🔊 Đang đọc...")

        # Chạy vòng đọc trên thread riêng để UI không bị đơ
        # daemon=True: thread tự tắt khi app đóng
        self.reading_thread = threading.Thread(
            target=self._reading_loop,
            daemon=True
        )
        self.reading_thread.start()

    def on_space(self, event=None):
        """
        Xử lý phím Space.
        Đặt cờ stop_requested = True → vòng đọc sẽ dừng sau dòng hiện tại.
        """
        if not self.is_reading:
            return
        self.stop_requested = True
        self.status_label.config(text="⏸ Sẽ dừng sau dòng này...")

    def _reading_loop(self):
        """
        Vòng lặp đọc — chạy trên thread riêng.
        Đọc từng dòng từ current_index đến hết,
        dừng lại nếu stop_requested = True.
        """
        while self.current_index < len(self.rows):

            # Kiểm tra cờ dừng trước mỗi dòng
            if self.stop_requested:
                break

            row = self.rows[self.current_index]

            # Cập nhật UI từ thread phụ phải dùng root.after
            # để tránh lỗi Tkinter (UI chỉ được cập nhật từ main thread)
            self.root.after(0, self._on_row_start, self.current_index)

            # Đọc dòng này — hàm này block cho đến khi đọc xong
            self.speak_row(row, self.headers)

            # Ghi X vào file Excel tại cột cuối + 1
            self._mark_done_in_excel(self.current_index)

            # Cập nhật UI: dòng vừa đọc xong → tô xanh
            self.root.after(0, self._on_row_done, self.current_index)

            # Sang dòng tiếp theo
            self.current_index += 1

        # Ra khỏi vòng lặp (đọc hết hoặc bị dừng)
        self.is_reading = False
        self.stop_requested = False

        # Cập nhật trạng thái trên UI
        if self.current_index >= len(self.rows):
            self.root.after(0, lambda: self.status_label.config(text="✅ Đã đọc hết file"))
        else:
            self.root.after(0, lambda: self.status_label.config(text="⏹ Đã dừng — nhấn Enter để tiếp tục"))

    def _on_row_start(self, index):
        """
        Gọi khi bắt đầu đọc một dòng.
        Chuyển trang nếu cần, render lại bảng để tô vàng dòng đang đọc.
        """
        self._go_to_page_of(index)
        self._render_page()
        self._update_progress()

        # Scroll đến dòng đang đọc trong trang hiện tại
        items = self.tree.get_children()
        row_in_page = index - self.page * self.page_size
        if 0 <= row_in_page < len(items):
            self.tree.see(items[row_in_page])

    def _on_row_done(self, index):
        """
        Gọi khi đọc xong một dòng.
        Render lại bảng để dòng đó chuyển sang màu xanh.
        """
        self._render_page()

    # ============================================================
    # PHẦN 7 — GHI X VÀO EXCEL
    # ============================================================

    def _mark_done_in_excel(self, index):
        """
        Ghi chữ X vào cột ngay sau cột cuối cùng của dòng đã đọc xong.
        index: index trong self.rows (bắt đầu từ 0)
        Dòng trong Excel = index + 2 (vì dòng 1 là header, index bắt đầu từ 0)
        """
        if self.workbook is None or self.filepath is None:
            return

        sheet = self.workbook.active
        excel_row = index + 2               # +1 vì header, +1 vì index từ 0
        last_col = len(self.headers) + 1    # Cột ngay sau cột cuối

        sheet.cell(row=excel_row, column=last_col, value="X")

        # Lưu file — ghi đè lên file gốc
        self.workbook.save(self.filepath)

    # ============================================================
    # PHẦN 8 — CẬP NHẬT TIẾN ĐỘ
    # ============================================================

    def _update_progress(self):
        """Cập nhật label tiến độ: dòng X / tổng Y."""
        if self.rows:
            self.progress_label.config(
                text=f"Dòng {self.current_index + 1} / {len(self.rows)}"
            )


# ============================================================
# PHẦN 9 — CHẠY APP
# ============================================================

if __name__ == "__main__":
    # Tạo cửa sổ Tkinter chính
    root = tk.Tk()

    # Tạo app — truyền root vào để app có thể điều khiển cửa sổ
    app = ExcelTTSApp(root)

    # mainloop() giữ app chạy, lắng nghe sự kiện (click, phím, ...)
    # cho đến khi người dùng đóng cửa sổ
    root.mainloop()